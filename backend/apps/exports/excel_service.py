import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_attendance_excel(event):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    ws.merge_cells('A1:G1')
    ws['A1'] = f'EventMaster — {event.title}'
    ws['A1'].font = Font(bold=True, size=14, color='1E3A5F')
    ws['A1'].alignment = center

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Date: {event.date.strftime('%d/%m/%Y %H:%M')} | Location: {event.location}"
    ws['A2'].font = Font(italic=True, size=10, color='6B7280')
    ws['A2'].alignment = center

    ws.append([])

    headers = ['N°', 'Nom complet', 'Email', 'Téléphone', 'Statut', 'Date inscription', 'Notes']
    ws.append(headers)
    header_row = ws.max_row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    regs = event.registrations.filter(status__in=['pending', 'confirmed']).select_related('participant')
    for i, reg in enumerate(regs, 1):
        p = reg.participant
        row_data = [
            i, p.name, p.email, p.phone or '-',
            reg.status.capitalize(),
            reg.registered_at.strftime('%d/%m/%Y %H:%M'),
            reg.notes or ''
        ]
        ws.append(row_data)
        data_row = ws.max_row
        fill = alt_fill if i % 2 == 0 else None
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=data_row, column=col)
            cell.alignment = center
            cell.border = thin_border
            if fill:
                cell.fill = fill

    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.row_dimensions[header_row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
